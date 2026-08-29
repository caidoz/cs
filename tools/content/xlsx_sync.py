# -*- coding: utf-8 -*-
"""content/*.tsv <-> content/*.xlsx

밸런스는 엑셀에서 잡고, 값은 tsv 를 거쳐 소스로 간다.

    엑셀  ->  tsv  ->  Data/*.cpp  ->  content.pack
          이 도구      content_table.py   build_pack.py

[왜 tsv 를 거치나]
.cpp 를 쓰는 도구는 content_table.py 하나여야 한다. 그 도구가 6666 칸을
한 칸씩 대조하는 verify 를 들고 있다. 엑셀에서 .cpp 로 바로 쓰면 그 검사를
건너뛰게 되고, 검사 없는 경로가 하나 생기면 없는 것과 같다.

[쓰는 법]
    python tools/content/xlsx_sync.py export          #tsv -> xlsx (전부)
    python tools/content/xlsx_sync.py export unit     #한 묶음만
    python tools/content/xlsx_sync.py import unit     #xlsx -> tsv
    python tools/content/xlsx_sync.py import unit --build
                                                      #tsv -> 소스 -> 팩까지

[정렬 금지]
엑셀에는 정렬 버튼이 있다. 그런데 이 표의 id 는 세이브 파일과 서버 DB 에
이미 나간 번호다(content/README.md). 한 줄이라도 자리가 바뀌면 남의
슬라임이 해골이 된다.

그래서 import 는 값을 쓰기 전에 세 가지를 본다.

    1. id 가 0 부터 하나씩 오르는가   <- 정렬을 잡는다
    2. 열 이름이 그대로인가            <- 열을 끼우거나 지운 것을 잡는다
    3. 줄 수가 그대로인가              <- 줄을 지운 것을 잡는다

하나라도 어긋나면 아무것도 쓰지 않고 멈춘다.
"""
import argparse
import io
import os
import subprocess
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, '..', '..'))
CONTENT = os.path.join(ROOT, 'content')

#---------------------------------------------------------------- 묶음
#
#한 파일이 질문 하나에 답하게 나눈다. "어느 .cpp 에 있느냐"가 아니라
#"한 번에 같이 잡는 것이냐"가 기준이다.
BOOKS = {
    'unit': {
        'title': '적과 동료와 스킬',
        'tables': ['enemy', 'crew', 'skill'],
        'note': '적이 얼마나 센가. 동료가 무엇을 하는가.',
    },
    'hero': {
        'title': '장비',
        'tables': ['item_pow', 'item_star', 'equip_value', 'sword'],
        'note': '로빈이 얼마나 세지는가. 시트마다 한 행의 뜻이 다르다 - item_pow 는 티어 하나, item_star 는 아이템 하나, equip_value 는 장비 한 점, sword 는 검 x 등급 한 칸이다.',
    },
    'reward': {
        'title': '보상',
        'note': '얼마나 받는가. castle 은 한 행이 성 하나이고, castle_box 는 한 행이 진행 순서 자리다 - 색인의 뜻이 달라서 시트를 나눴다.',
        'tables': ['castle', 'castle_box', 'reward_levelup', 'reward_bossgold',
                   'reward_battlegold', 'reward_stageclear', 'reward_boss',
                   'reward_crew', 'reward_equipheart'],
    },
    'economy': {
        'title': '재화와 상점',
        'note': '얼마에 파는가. 베팅 배수, 골드 곡선, 아이템 값, 상점.',
        'tables': ['bet_heart', 'bet_coin', 'gold_stage', 'gold_skill',
                   'gold_crewstar', 'item_price', 'item_sell',
                   'shop_currency', 'shop_quick'],
    },
    'wave': {
        'title': '웨이브',
        'note': '적을 언제 만나는가. wave[] 90000 칸은 여기 규칙으로 make_waves.py 가 깐다 - 결과가 아니라 손잡이를 고치는 곳이다. 규칙으로 안 되는 자리는 wave_fix 에 한 줄씩 적으면 생성기를 이긴다.',
        'tables': ['wave_rule', 'wave_region', 'wave_color', 'wave_fix'],
        'grow': ['wave_fix'],
        'build': 'make_waves.py',
    },
}

#열 설명. 없는 열은 이름만 나온다.
DESC = {
    'skill': {
        'kind': '종류. 이 값이 나머지 칸의 뜻을 정한다',
        'value': '효능 수치 (히어로 스킬의 버프 % 등)',
        'hit_max': '한 번에 몇 대까지. 1 이면 단발',
        'dmg_pct': '한 대가 기본공격의 몇 %. 100 이 그대로, 300 이 3 배',
        'cooldown': '재사용 대기 (히어로 스킬)',
        'icon': '아이콘 번호',
        'icon_kind': '아이콘이 어느 그림판에서 오는가',
        'star': '카드의 별',
        'attack_type': '어떤 공격으로 나가는가',
        'host_obj': '어느 오브젝트 칸에서 벌어지는가',
        'summon_enemy': '불러낼 몬스터 (SUMMON)',
        'hero_type': '불러낼 히어로 (SUMMONHERO)',
        'hero_skill': '발동시킬 히어로 스킬',
        'summon_x': '소환체가 설 x. 0 이면 화면 중앙',
        'bullet_obj': '총알 오브젝트 (CREWBULLET)',
        'bullet_icon': '총알 그림 (CREWBULLET)',
        'equip1': '소환 히어로에게 입힐 장비 (SUMMONHERO)',
    },
    'crew': {
        'enemy_type': '어느 몬스터 그림을 쓰는가',
        'str': '기본 공격력. 여기에 별 x 레벨 배율이 곱해진다',
        'skill1': '슬롯에 1 개 떴을 때',
        'skill2': '슬롯에 2 개 떴을 때',
        'skill3': '슬롯에 3 개 떴을 때',
        'card_bg': '카드 뒷그림',
    },
    'castle': {'star_limit': '이 성에서 나올 수 있는 최대 별',
               'order': '이 성이 진행 순서 몇 번째인가'},
    'castle_box': {'box_gold': '그 자리에서 주는 골드',
                   'box_color': '상자 색'},
    'bet_heart': {'mul': '베팅 단계별 배수. 데미지에 곱해진다'},
    'gold_stage': {'gold': '스테이지 * TOTALROOM + room 으로 찾는다'},
    'wave_rule': {'name': '손잡이 이름. 고치지 말 것',
                  'value': '이 값을 고치고 다시 깐다',
                  'note': '무엇을 정하는 값인가'},
    'wave_region': {'order': '진행 순서. 앞이 약하다',
                    'castle': 'CastleDef.h 의 이름. 함부로 고치면 안 된다'},
    'wave_color': {'order': '단계. 0 이 기본색이다',
                   'suffix': '몬스터 이름에 붙는 꼬리. 빈 줄이 기본색'},
    'wave_fix': {'wave': '몇 번째 웨이브인가 (1 부터)',
                 'slot': '그 웨이브의 몇 번째 자리인가 (1~3)',
                 'enemy': '몬스터 열거 이름. 색까지 적는다',
                 'kind': 'MONSTERTYPE_JACO / _BOSS / _MIDBOSS / _BIGBOSS',
                 'note': '왜 손으로 잡았는지. 안 적으면 나중에 아무도 모른다'},
    'item_pow': {'pow': '그 티어의 기본 위력. GetItemPow 가 본다'},
    'item_star': {'star': '별 x 100. 600 이면 6 성이다'},
    'equip_value': {'value': '장비 한 점의 기본값. 강화 배율이 여기 곱해진다'},
    'sword': {'max_bet': '최대 베팅', 'heart': '하트 소모', 'gold': '골드'},
    'enemy': {
        'cmf': '그림 묶음',
        'star': '등급',
        'stat_hp': '체력',
        'stat_str': '공격력',
        'stat_def': '방어력. 데미지에서 절대값으로 빠진다',
        'add_exp': '잡았을 때 주는 경험치',
        'home_castle': '어느 지역 출신인가',
    },
}

#칸 이름에 이 말이 들어 있으면 그 색으로 묶어 보여준다
GROUPS = [
    (('id', 'enum_name', 'label'), 'FF1B365D'),      #이름표. 고치면 안 되는 것
    (('star', 'grade', 'kind'), 'FF8E44AD'),         #등급/종류
    (('str', 'stat_', 'value', 'dmg_', 'hit_'), 'FFC0392B'),   #세기
    (('icon', 'cmf', 'pose_', 'zoom', 'bigicon', 'skillicon'), 'FF16A085'),  #그림
    (('skill', 'hero_', 'summon', 'bullet', 'equip'), 'FFD35400'),           #참조
]
DEFCOLOR = 'FF7F8C8D'

FONT_HEAD = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFFFF')
FONT_ID = Font(name='Consolas', size=10, bold=True, color='FF7F8C8D')
FONT_CELL = Font(name='Consolas', size=10)
FONT_NAME = Font(name='맑은 고딕', size=10)
THIN = Side(style='thin', color='FFD0D3D4')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_LOCK = PatternFill('solid', start_color='FFF2F3F4', end_color='FFF2F3F4')


def color_of(col):
    for keys, c in GROUPS:
        for k in keys:
            if col == k or col.startswith(k):
                return c
    return DEFCOLOR


def tsv_path(name):
    return os.path.join(CONTENT, name + '.tsv')


def xlsx_path(book):
    return os.path.join(CONTENT, book + '.xlsx')


def read_tsv(name):
    """(주석줄, 열이름, 행들)"""
    with io.open(tsv_path(name), encoding='utf-8', newline='') as fp:
        raw = fp.read()

    nl = '\r\n' if '\r\n' in raw else '\n'
    lines = raw.split(nl)
    head = [l for l in lines if l.startswith('#')]
    body = [l for l in lines if l and not l.startswith('#')]

    return head, body[0].split('\t'), [l.split('\t') for l in body[1:]], nl


def write_tsv(name, head, cols, rows, nl):
    out = head + ['\t'.join(cols)] + ['\t'.join(r) for r in rows]

    with io.open(tsv_path(name), 'w', encoding='utf-8', newline='') as fp:
        fp.write(nl.join(out) + nl)



#---------------------------------------------------------------- 그림 열
#
#숫자만 보고 밸런스를 잡기는 어렵다. 그래서 시트에 그림을 한 열 끼운다.
#
#이 열은 tsv 에 없다. 보기만 하는 열이라 import 가 읽지 않고 버린다.
#엑셀에서 그림은 셀 값이 아니라 시트에 떠 있는 개체라, 값만 읽는 import
#입장에서는 애초에 보이지도 않는다. 머리글만 이름으로 걸러 내면 된다.
#
#  (끼울 자리, 머리글, 그림을 고르는 함수)
#자리는 0 부터 세는 tsv 열 번호다. 3 이면 label 다음이다.

ICON_DIR = os.path.join(CONTENT, 'icon')


def skill_icon(col, row):
    """스킬 한 줄이 어느 그림을 쓰는가. 없으면 None."""
    kind = row[col['icon_kind']]
    idx = row[col['icon']]

    sheet = {'ICONKIND_SKILL': 'skill', 'ICONKIND_BULLET': 'bullet'}.get(kind)

    if sheet is None:
        return None            #몬스터는 게임이 그려 줘야 한다(2 단계)

    p = os.path.join(ICON_DIR, '%s_%s.png' % (sheet, idx))

    return p if os.path.isfile(p) else None


DISPLAY = {
    'skill': [(3, '그림', skill_icon)],
}

IMG_PX = 32                     #그림 한 변
IMG_ROW_H = 26                  #그림이 들어가는 줄 높이(포인트)
IMG_COL_W = 6


def display_of(name):
    return DISPLAY.get(name, [])


def header_with_display(name, cols):
    """tsv 열 이름에 그림 열을 끼운 머리글과, 끼운 자리 목록."""
    out = list(cols)
    at = []

    for pos, title, _fn in sorted(display_of(name), reverse=True):
        out.insert(pos, title)

    for i, h in enumerate(out):
        if any(h == t for _p, t, _f in display_of(name)):
            at.append(i)

    return out, at


#---------------------------------------------------------------- 내보내기

def sheet_from(wb, name):
    head, cols, rows, _ = read_tsv(name)
    ws = wb.create_sheet(name)

    full, dispAt = header_with_display(name, cols)

    #tsv 열이 시트의 몇 번째 칸으로 가는가(1 부터)
    colAt = []
    k = 0

    for i in range(len(full)):
        if i in dispAt:
            continue

        colAt.append(i + 1)
        k += 1

    for c, col in enumerate(full, start=1):
        cell = ws.cell(1, c, col)
        cell.font = FONT_HEAD
        cell.fill = PatternFill('solid', start_color=color_of(col),
                                end_color=color_of(col))
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border = BORDER

        d = DESC.get(name, {}).get(col)

        if d:
            cell.comment = openpyxl.comments.Comment(d, 'xlsx_sync')

    for r, row in enumerate(rows, start=2):
        for i, v in enumerate(row):
            c = colAt[i]
            cell = ws.cell(r, c, num(v))
            cell.border = BORDER

            if c <= 3:
                cell.font = FONT_ID if c == 1 else FONT_NAME
                cell.fill = FILL_LOCK
            else:
                cell.font = FONT_CELL

    put_images(ws, name, cols, rows, dispAt)

    #머리글과 이름 세 칸을 붙들어 둔다
    ws.freeze_panes = 'D2'
    ws.auto_filter.ref = None       #필터를 달면 정렬을 부른다. 달지 않는다

    #kind 같은 칸은 지금 표에 있는 값만 고르게 한다
    for i, col in enumerate(cols):
        vals = sorted(set(r[i] for r in rows))

        if col in ('kind', 'icon_kind') and 1 < len(vals) <= 20:
            L = get_column_letter(colAt[i])
            dv = DataValidation(type='list',
                                formula1='"%s"' % ','.join(vals),
                                allow_blank=False)
            ws.add_data_validation(dv)
            dv.add('%s2:%s%d' % (L, L, len(rows) + 1))

    #열 너비
    for i, col in enumerate(cols):
        w = max(len(col) + 2,
                max((len(str(r[i])) for r in rows), default=4) + 2)
        ws.column_dimensions[get_column_letter(colAt[i])].width = min(w, 30)

    for i in dispAt:
        ws.column_dimensions[get_column_letter(i + 1)].width = IMG_COL_W

    ws.row_dimensions[1].height = 30

    return len(rows), len(cols)


def put_images(ws, name, cols, rows, dispAt):
    """그림 열에 낱장 PNG 를 얹는다. 파일이 없으면 그냥 빈 칸이다."""
    spec = display_of(name)

    if not spec:
        return

    col = {c: i for i, c in enumerate(cols)}
    n = 0

    for (pos, _title, fn), at in zip(sorted(spec), dispAt):
        letter = get_column_letter(at + 1)

        for r, row in enumerate(rows, start=2):
            p = fn(col, row)

            if not p:
                continue

            img = XLImage(p)
            img.width = IMG_PX
            img.height = IMG_PX
            ws.add_image(img, '%s%d' % (letter, r))
            ws.row_dimensions[r].height = IMG_ROW_H
            n += 1

    if n:
        print('    그림 %d개' % n)


def guide(wb, book):
    ws = wb.create_sheet('읽기', 0)
    L = [
        (BOOKS[book]['title'], True),
        (BOOKS[book]['note'], False),
        ('', False),
        ('[고치는 법]', True),
        ('1. 값을 고친다', False),
        ('2. 저장하고 엑셀을 닫는다', False),
        ('3. python tools/content/xlsx_sync.py import %s --build' % book, False),
        ('', False),
        ('[하면 안 되는 것]', True),
        ('줄을 정렬하지 말 것. id 는 세이브와 서버에 이미 나간 번호라,',
         False),
        ('자리가 바뀌면 남의 슬라임이 해골이 된다.', False),
        ('줄이나 열을 지우거나 끼우지도 말 것.', False),
        ('', False),
        ('import 가 id 순서와 열 이름과 줄 수를 보고, 어긋나면 아무것도',
         False),
        ('쓰지 않고 멈춘다. 그래도 되돌리는 것보다 안 하는 편이 낫다.',
         False),
        ('', False),
        ('[새 줄을 넣으려면]', True),
        ('맨 끝에만 붙인다. 그리고 Def/*.h 의 열거에도 같은 이름을 맨',
         False),
        ('끝에 넣어야 한다. 그건 아직 손으로 한다.', False),
        ('', False),
        ('[열 색]', True),
        ('남색 = 이름표(고치지 말 것)   보라 = 등급/종류', False),
        ('빨강 = 세기                   초록 = 그림', False),
        ('주황 = 다른 표를 가리키는 값', False),
    ]

    for i, (t, b) in enumerate(L, start=1):
        cell = ws.cell(i, 1, t)
        cell.font = Font(name='맑은 고딕', size=11, bold=b)

    ws.column_dimensions['A'].width = 70


def num(v):
    """숫자로 보이면 숫자로. 엑셀에서 계산할 수 있게."""
    try:
        return int(v)
    except (TypeError, ValueError):
        return v


def do_export(book):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for t in BOOKS[book]['tables']:
        r, c = sheet_from(wb, t)
        print('  %-8s %4d행 x %2d열' % (t, r, c))

    guide(wb, book)
    p = xlsx_path(book)

    try:
        wb.save(p)
    except PermissionError:
        sys.exit('%s 를 못 쓴다. 엑셀에서 열려 있으면 닫아라' % p)

    print('%s' % p)


#---------------------------------------------------------------- 들여오기

def do_import(book, build):
    p = xlsx_path(book)

    if not os.path.isfile(p):
        sys.exit('%s 가 없다. 먼저 export 해라' % p)

    wb = openpyxl.load_workbook(p, data_only=True)
    plan = []

    for name in BOOKS[book]['tables']:
        if name not in wb.sheetnames:
            sys.exit('%s 시트가 없다' % name)

        ws = wb[name]
        head, cols, rows, nl = read_tsv(name)

        full, dispAt = header_with_display(name, cols)
        got = [ws.cell(1, c).value for c in range(1, len(full) + 1)]

        #그림 열은 보기만 하는 열이다. 버리고 나머지를 표와 맞춘다.
        got = [v for i, v in enumerate(got) if i not in dispAt]

        if got != cols:
            sys.exit('%s : 열 이름이 다르다\n  표 %s\n  엑셀 %s'
                     % (name, cols, got))

        grow = name in BOOKS[book].get('grow', [])

        if not grow and ws.max_row - 1 != len(rows):
            sys.exit('%s : 줄이 %d개여야 하는데 %d개다'
                     % (name, len(rows), ws.max_row - 1))

        new = []

        for r in range(2, (ws.max_row + 1) if grow else (len(rows) + 2)):
            vals = []

            for c in range(1, len(full) + 1):
                if (c - 1) in dispAt:
                    continue

                v = ws.cell(r, c).value
                vals.append('' if v is None else str(v).strip())

            #통째로 빈 줄은 버린다. 엑셀은 지운 줄을 빈 줄로 남긴다.
            if grow and not any(v for v in vals):
                continue

            #id 가 제자리인가. 정렬을 여기서 잡는다.
            #늘어나도 되는 표는 있던 줄까지만 본다.
            if r - 2 < len(rows) and vals[0] != rows[r - 2][0]:
                sys.exit('%s : %d번째 줄의 id 가 %s 여야 하는데 %s 다.\n'
                         '정렬했거나 줄을 옮긴 것 같다. 아무것도 쓰지 않았다.'
                         % (name, r - 1, rows[r - 2][0], vals[0]))

            new.append(vals)

        changed = sum(1 for a, b in zip(rows, new) for x, y in zip(a, b)
                      if x != y)
        plan.append((name, head, cols, new, nl, changed))

    for name, head, cols, new, nl, changed in plan:
        write_tsv(name, head, cols, new, nl)
        print('  %-8s %d칸 바뀜' % (name, changed))

    if build:
        #표마다 소스를 내는 도구가 다르다. 웨이브는 규칙에서 까는 것이라
        #content_table 이 아니라 make_waves 가 낸다.
        tool = BOOKS[book].get('build')

        if tool:
            subprocess.check_call([sys.executable, os.path.join(HERE, tool)])
        else:
            for name in BOOKS[book]['tables']:
                subprocess.check_call([sys.executable,
                                       os.path.join(HERE, 'content_table.py'),
                                       name, 'generate', '--write'])

        subprocess.check_call([sys.executable,
                               os.path.join(HERE, 'build_pack.py')])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('what', choices=['export', 'import'])
    ap.add_argument('book', nargs='?', default=None,
                    help='묶음 이름. 없으면 전부')
    ap.add_argument('--build', action='store_true',
                    help='import 뒤에 소스와 팩까지 다시 낸다')
    a = ap.parse_args()

    books = [a.book] if a.book else sorted(BOOKS)

    for b in books:
        if b not in BOOKS:
            sys.exit('%s 는 없는 묶음이다. 있는 것 : %s'
                     % (b, ', '.join(sorted(BOOKS))))

        print('[%s]' % b)

        if a.what == 'export':
            do_export(b)
        else:
            do_import(b, a.build)


main()
